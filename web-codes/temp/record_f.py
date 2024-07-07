import mysql.connector
import xlsxwriter
import sys
from datetime import datetime
import to_excel
import openpyxl

mydb = mysql.connector.connect(host = "localhost", user = "pi", passwd = "pi", database = "Attendance")
mycursor = mydb.cursor() 
mycursor.execute("DROP TABLE IF EXISTS tmp")
year = int(sys.argv[2])
if (sys.argv[1] in ["01","03","05","07","08","10","12"]):
	day = 31
elif(sys.argv[1] in ["04","06","09","11"]):
	day = 30
else:
	if (year % 4) == 0:
   		if (year % 100) == 0:
      			if (year % 400) == 0:
          			day=29
      			else:
           			day=28
   		else:
       			day=29
	else:
   		day=28

dc=str(datetime.now().date())

if (int(year) == int(dc[-10:-6]) and int(sys.argv[1]) == int(dc[-5:-3])):
	day=int(dc[-2:])-1
	 
d1,d2=str(year)+"-"+sys.argv[1]+"-01",str(year)+"-"+sys.argv[1]+"-"+str(day)


mycursor.execute("DROP TABLE IF EXISTS tmp")
sql ="CREATE TABLE tmp(NAME CHAR(100) NOT NULL,"
for i in range(1,day+1):
	
	sql=sql+"day_"+str(i)+" CHAR(100) NOT NULL,"
sql= sql[0:-1]
sql = sql =sql
sql=sql+")"



#if day==30:
#	sql ='CREATE TABLE tmp(NAME CHAR(100) NOT NULL,date1 CHAR(100),date2 CHAR(100),date3 CHAR(100),date4 CHAR(100),date5 CHAR(100),date6 CHAR(100),date7 CHAR(100),date8 CHAR(100),date9 CHAR(100),date10 CHAR(100),date11 CHAR(100),date12 CHAR(100),date13 CHAR(100),date14 CHAR(100),date15 CHAR(100),date16 CHAR(100),date17 CHAR(100),date18 CHAR(100),date19 CHAR(100),date20 CHAR(100),date21 CHAR(100),date22 CHAR(100),date23 CHAR(100),date24 CHAR(100),date25 CHAR(100),date26 CHAR(100),date27 CHAR(100),date28 CHAR(100),date29 CHAR(100),date30 CHAR(100))'

#elif day==31:
#	sql ='CREATE TABLE tmp(NAME CHAR(100) NOT NULL,date1 CHAR(100),date2 CHAR(100),date3 CHAR(100),date4 CHAR(100),date5 CHAR(100),date6 CHAR(100),date7 CHAR(100),date8 CHAR(100),date9 CHAR(100),date10 CHAR(100),date11 CHAR(100),date12 CHAR(100),date13 CHAR(100),date14 CHAR(100),date15 CHAR(100),date16 CHAR(100),date17 CHAR(100),date18 CHAR(100),date19 CHAR(100),date20 CHAR(100),date21 CHAR(100),date22 CHAR(100),date23 CHAR(100),date24 CHAR(100),date25 CHAR(100),date26 CHAR(100),date27 CHAR(100),date28 CHAR(100),date29 CHAR(100),date30 CHAR(100),date31 CHAR(100))'

#elif day==29:
#	sql ='CREATE TABLE tmp(NAME CHAR(100) NOT NULL,date1 CHAR(100),date2 CHAR(100),date3 CHAR(100),date4 CHAR(100),date5 CHAR(100),date6 CHAR(100),date7 CHAR(100),date8 CHAR(100),date9 CHAR(100),date10 CHAR(100),date11 CHAR(100),date12 CHAR(100),date13 CHAR(100),date14 CHAR(100),date15 CHAR(100),date16 CHAR(100),date17 CHAR(100),date18 CHAR(100),date19 CHAR(100),date20 CHAR(100),date21 CHAR(100),date22 CHAR(100),date23 CHAR(100),date24 CHAR(100),date25 CHAR(100),date26 CHAR(100),date27 CHAR(100),date28 CHAR(100),date29 CHAR(100))'
#else:
#	sql ='CREATE TABLE tmp(NAME CHAR(100) NOT NULL,date1 CHAR(100),date2 CHAR(100),date3 CHAR(100),date4 CHAR(100),date5 CHAR(100),date6 CHAR(100),date7 CHAR(100),date8 CHAR(100),date9 CHAR(100),date10 CHAR(100),date11 CHAR(100),date12 CHAR(100),date13 CHAR(100),date14 CHAR(100),date15 CHAR(100),date16 CHAR(100),date17 CHAR(100),date18 CHAR(100),date19 CHAR(100),date20 CHAR(100),date21 CHAR(100),date22 CHAR(100),date23 CHAR(100),date24 CHAR(100),date25 CHAR(100),date26 CHAR(100),date27 CHAR(100),date28 CHAR(100))'

mycursor.execute(sql)

if(sys.argv[3]=="student"):
	heading=sys.argv[3]+" records : "+d1+" to "+d2+" ----Class "+sys.argv[4]+" ----"
	sql="select fingerprintid,name from registered_members where category='student' and class ="+sys.argv[4]
else:
	heading=sys.argv[3]+" records : "+d1+" to "+d2
	sql="select fingerprintid,name from registered_members where category ='"+sys.argv[3]+"'"
mycursor.execute(sql)
datatupfin=()
datatupname=()
if(mycursor.rowcount==0):
	#print ("No row found")	
	exit(0)
for x in mycursor:
	datatupfin=datatupfin+(str(x[0]),)
	datatupname=datatupname+(str(x[1]),)


for y,yn in zip(datatupfin,datatupname):
	datetup=()
	sql="select date from member_attendance where shift =2 and fingerprintid= "+y+" and date between '"+d1+"' and '"+d2+"'"

	mycursor.execute(sql)
	if(mycursor.rowcount==0):
		#print "No row found"	
		exit(0)
	for z in mycursor:
		datetup=datetup+(int(str(z[0])[-2:]),)
	
	tupgo = (yn,)
	for i in range(1,day+1):
		
		if(i in datetup):
			tupgo = tupgo+("P",)
		else:
			tupgo = tupgo+("A",)

	
	
	sql="insert into tmp values"+str(tupgo)
	mycursor.execute(sql)
	mydb.commit()
	
	
	del tupgo
	
	del datetup


mycursor.close()
mydb.close()
to_excel.export('tmp','SchoolReport')
xfile = openpyxl.load_workbook('SchoolReport.xlsx')

sheet = xfile.get_sheet_by_name('record')

sheet.insert_rows(1)
sheet.insert_rows(1)
sheet.merge_cells('A1:AF1')
sheet.merge_cells('A2:AF2')
sheet['A1'] = "           "+heading
xfile.save('SchoolReport.xlsx')
